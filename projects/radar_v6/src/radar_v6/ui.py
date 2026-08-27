import json
import sys
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from .data import (
    list_sheet_names,
    load_people_scores,
    normalize_export_name,
    normalize_search_text,
    normalize_text,
)
from .mapping import (
    looks_like_group_header,
    looks_like_subtotal,
    parse_col_ref,
    pick_recommended_header,
    sanitize_template_filename,
    validate_mapping_config,
)
from .export import draw_radar


def get_app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path.cwd().resolve()

BASE_DIR = get_app_base_dir()
SOURCE_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL_NAME = "??3 ?????????.xlsx"
DEFAULT_EXCEL_PATH = next(
    (
        p
        for p in (
            BASE_DIR / DEFAULT_EXCEL_NAME,
            BASE_DIR.parent / DEFAULT_EXCEL_NAME,
            SOURCE_DIR / DEFAULT_EXCEL_NAME,
        )
        if p.exists()
    ),
    BASE_DIR / DEFAULT_EXCEL_NAME,
)
DEFAULT_SHEET_NAME = "Sheet1"
APP_SETTINGS_FILE = "radar_v6_settings.json"
DEFAULT_TEMPLATE_FILE = "????.json"


class RadarApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("员工能力雷达图 V6")
        self.geometry("1320x820")
        self.minsize(1080, 680)

        self.default_mapping_raw = {
            "name": "默认模板",
            "name_col": 2,
            "data_start_row": 6,
            "header_row_primary": 4,
            "header_row_secondary": 5,
            "score_max": 10,
            "dimensions": [
                {"name": "身体素质", "direct_col": 11, "sum_cols": [8, 9, 10], "item_names": ["年龄", "身体健康", "心理健康"]},
                {"name": "资质资历", "direct_col": 16, "sum_cols": [12, 13, 14, 15], "item_names": ["学历", "职称", "本岗位工作年限", "具备资格证件"]},
                {"name": "工作能力", "direct_col": 17, "sum_cols": [17], "item_names": []},
                {"name": "QHSE能力", "direct_col": 22, "sum_cols": [18, 19, 20, 21], "item_names": ["QHSE意愿", "岗位基本知识", "岗位QHSE技能", "应急处置能力"]},
                {"name": "工作态度", "direct_col": 27, "sum_cols": [23, 24, 25, 26], "item_names": ["责任心", "执行力", "主动性", "协作精神"]},
            ],
        }
        self.mapping_config_raw = json.loads(json.dumps(self.default_mapping_raw, ensure_ascii=False))
        self.mapping_config = validate_mapping_config(self.mapping_config_raw)
        self.labels = [d["name"] for d in self.mapping_config["dimensions"]]
        self.people = []
        self.people_by_name = {}
        self.filtered_names = []
        self.dimension_text_map = {}
        self.current_person = None
        self.current_dimension = self.labels[0] if self.labels else ""
        self.export_current_button = None
        self.export_batch_button = None
        self.dim_btn_frame = None

        self.excel_path_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.template_name_var = tk.StringVar(value=f"模板：{self.mapping_config_raw['name']}")
        self.template_file_var = tk.StringVar(value="模板文件：")

        self.settings_path = BASE_DIR / APP_SETTINGS_FILE
        self.settings = self.load_app_settings()
        self.template_dir = Path(self.settings.get("template_dir", str(BASE_DIR)))
        current_path = str(self.settings.get("current_template_path", "")).strip()
        self.current_template_path = Path(current_path) if current_path else None

        self.columnconfigure(0, weight=0)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        left = ttk.Frame(self, padding=12)
        left.grid(row=0, column=0, sticky="ns")
        left.rowconfigure(8, weight=1)

        ttk.Label(left, text="数据文件", font=("Microsoft YaHei", 11, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )
        path_row = ttk.Frame(left)
        path_row.grid(row=1, column=0, sticky="ew", pady=(0, 6))
        path_row.columnconfigure(0, weight=1)
        self.path_entry = ttk.Entry(path_row, textvariable=self.excel_path_var, width=28, state="readonly")
        self.path_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(path_row, text="选择Excel", command=self.on_choose_excel).grid(row=0, column=1, sticky="e")

        sheet_row = ttk.Frame(left)
        sheet_row.grid(row=2, column=0, sticky="ew", pady=(0, 6))
        sheet_row.columnconfigure(0, weight=1)
        self.sheet_combo = ttk.Combobox(
            sheet_row,
            textvariable=self.sheet_var,
            state="readonly",
            width=18,
        )
        self.sheet_combo.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_change)
        ttk.Button(sheet_row, text="重新加载", command=self.reload_people).grid(row=0, column=1, sticky="e")

        ttk.Label(left, textvariable=self.template_name_var, foreground="#0A6FAE").grid(
            row=3, column=0, sticky="w", pady=(2, 6)
        )
        ttk.Label(left, textvariable=self.template_file_var, foreground="#666").grid(
            row=4, column=0, sticky="w", pady=(0, 6)
        )

        cfg_row = ttk.Frame(left)
        cfg_row.grid(row=5, column=0, sticky="ew", pady=(0, 10))
        ttk.Button(cfg_row, text="配置映射", command=self.open_mapping_editor).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(cfg_row, text="模板路径", command=self.choose_template_dir).grid(row=0, column=1, padx=(0, 6))
        ttk.Button(cfg_row, text="保存模板", command=self.save_template).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(cfg_row, text="加载模板", command=self.load_template).grid(row=0, column=3)

        ttk.Label(left, text="选择员工", font=("Microsoft YaHei", 12, "bold")).grid(
            row=6, column=0, sticky="w", pady=(0, 8)
        )

        self.search_var = tk.StringVar()
        search = ttk.Entry(left, textvariable=self.search_var, width=24)
        search.grid(row=7, column=0, sticky="ew", pady=(0, 8))
        search.bind("<KeyRelease>", self.on_search)

        list_frame = ttk.Frame(left)
        list_frame.grid(row=8, column=0, sticky="nsew")
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        self.listbox = tk.Listbox(list_frame, width=24, font=("Microsoft YaHei", 11))
        self.listbox.grid(row=0, column=0, sticky="nsew")
        self.listbox.bind("<<ListboxSelect>>", self.on_select)

        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.listbox.config(yscrollcommand=scrollbar.set)

        right = ttk.Frame(self, padding=(4, 8, 10, 8))
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(0, weight=1)
        right.rowconfigure(1, weight=0)
        right.columnconfigure(0, weight=1)

        self.fig = plt.Figure(
            figsize=(7.2, 7.2),
            dpi=100,
            facecolor="#FAFBFC",
            constrained_layout=True,
        )
        self.ax = self.fig.add_subplot(111, polar=True)
        self.fig.set_constrained_layout_pads(w_pad=0.03, h_pad=0.03, hspace=0.02, wspace=0.02)

        self.canvas = FigureCanvasTkAgg(self.fig, master=right)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.grid(row=0, column=0, sticky="nsew")
        self.canvas.mpl_connect("pick_event", self.on_chart_pick)

        action_row = ttk.Frame(right)
        action_row.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        action_row.columnconfigure(0, weight=1)
        self.export_batch_button = ttk.Button(
            action_row,
            text="批量导出",
            command=self.export_multiple_people_charts,
            state="disabled",
        )
        self.export_batch_button.grid(row=0, column=1, sticky="e", padx=(0, 6))
        self.export_current_button = ttk.Button(
            action_row,
            text="导出当前图像",
            command=self.export_current_chart,
            state="disabled",
        )
        self.export_current_button.grid(row=0, column=2, sticky="e")

        detail_frame = ttk.LabelFrame(right, text="维度明细（点击雷达图维度名称切换）", padding=10)
        detail_frame.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        detail_frame.columnconfigure(0, weight=1)

        self.detail_title_var = tk.StringVar(value=f"当前维度：{self.current_dimension}" if self.current_dimension else "当前维度：")
        ttk.Label(detail_frame, textvariable=self.detail_title_var, font=("Microsoft YaHei", 11, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )

        self.dim_btn_frame = ttk.Frame(detail_frame)
        self.dim_btn_frame.grid(row=1, column=0, sticky="w", pady=(0, 8))
        self.refresh_dimension_buttons()

        self.detail_table = ttk.Treeview(detail_frame, columns=("item", "value"), show="headings", height=6)
        self.detail_table.heading("item", text="指标")
        self.detail_table.heading("value", text="分数")
        self.detail_table.column("item", width=300, anchor="w")
        self.detail_table.column("value", width=120, anchor="center")
        self.detail_table.grid(row=2, column=0, sticky="ew")

        self.initialize_template_file()

        if DEFAULT_EXCEL_PATH.exists():
            self.set_excel_file(DEFAULT_EXCEL_PATH)

    def refresh_dimension_buttons(self):
        for child in self.dim_btn_frame.winfo_children():
            child.destroy()
        for idx, dim_name in enumerate(self.labels):
            ttk.Button(
                self.dim_btn_frame,
                text=dim_name,
                command=lambda n=dim_name: self.set_dimension(n),
                width=10,
            ).grid(row=0, column=idx, padx=(0, 6))

    def load_app_settings(self):
        if not self.settings_path.exists():
            return {}
        try:
            with open(self.settings_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def save_app_settings(self):
        data = {
            "template_dir": str(self.template_dir),
            "current_template_path": str(self.current_template_path) if self.current_template_path else "",
        }
        with open(self.settings_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def get_default_template_path(self):
        return self.template_dir / DEFAULT_TEMPLATE_FILE

    def write_template_file(self, path: Path, raw_config: dict):
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(raw_config, f, ensure_ascii=False, indent=2)

    def initialize_template_file(self):
        self.template_dir.mkdir(parents=True, exist_ok=True)
        default_path = self.get_default_template_path()
        if not default_path.exists():
            self.write_template_file(default_path, self.default_mapping_raw)

        target = self.current_template_path if self.current_template_path and self.current_template_path.exists() else default_path
        try:
            with open(target, "r", encoding="utf-8") as f:
                raw = json.load(f)
            self.apply_mapping_config(raw)
            self.current_template_path = target
        except Exception:
            self.apply_mapping_config(self.default_mapping_raw)
            self.current_template_path = default_path
            self.write_template_file(default_path, self.mapping_config_raw)

        self.template_file_var.set(f"模板文件：{self.current_template_path}")
        self.save_app_settings()

    def autosave_current_template(self):
        if not self.current_template_path:
            filename = f"{sanitize_template_filename(self.mapping_config_raw.get('name', '当前模板'))}.json"
            self.current_template_path = self.template_dir / filename
        if self.current_template_path.name == DEFAULT_TEMPLATE_FILE:
            filename = f"{sanitize_template_filename(self.mapping_config_raw.get('name', '当前模板'))}.json"
            if filename == DEFAULT_TEMPLATE_FILE:
                filename = "当前模板.json"
            self.current_template_path = self.template_dir / filename
        self.write_template_file(self.current_template_path, self.mapping_config_raw)
        self.template_file_var.set(f"模板文件：{self.current_template_path}")
        self.save_app_settings()

    def choose_template_dir(self):
        chosen = filedialog.askdirectory(title="选择模板保存目录", initialdir=str(self.template_dir))
        if not chosen:
            return
        self.template_dir = Path(chosen)
        self.template_dir.mkdir(parents=True, exist_ok=True)
        default_path = self.get_default_template_path()
        if not default_path.exists():
            self.write_template_file(default_path, self.default_mapping_raw)

        current_name = sanitize_template_filename(self.mapping_config_raw.get("name", "模板"))
        self.current_template_path = self.template_dir / f"{current_name}.json"
        self.autosave_current_template()
        messagebox.showinfo("已应用", f"模板目录已切换到：\n{self.template_dir}\n当前模板已自动保存。")

    def apply_mapping_config(self, raw_config: dict):
        parsed = validate_mapping_config(raw_config)
        self.mapping_config_raw = raw_config
        self.mapping_config = parsed
        self.labels = [d["name"] for d in parsed["dimensions"]]
        if self.current_dimension not in self.labels:
            self.current_dimension = self.labels[0] if self.labels else ""
        self.template_name_var.set(f"模板：{self.mapping_config_raw.get('name', '自定义模板')}")
        self.refresh_dimension_buttons()

    def split_csv(self, text: str):
        return [x.strip() for x in str(text).replace("，", ",").split(",") if x.strip()]

    def _forward_fill_headers(self, values):
        filled = []
        current = ""
        for val in values:
            txt = normalize_text(val)
            if txt:
                current = txt
            filled.append(current)
        return filled

    def detect_group_mapping_by_column(self, header_row_primary: int, header_row_secondary: int, picked_col_ref: str):
        excel_path_text = self.excel_path_var.get().strip()
        sheet_name = self.sheet_var.get().strip()
        excel_path = Path(excel_path_text)
        if not excel_path.exists():
            raise FileNotFoundError(f"找不到表格文件: {excel_path}")

        picked_col = parse_col_ref(picked_col_ref)
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        max_col = min(ws.max_column, 200)

        p_raw = [ws.cell(row=header_row_primary, column=c).value for c in range(1, max_col + 1)]
        s_raw = [ws.cell(row=header_row_secondary, column=c).value for c in range(1, max_col + 1)]
        wb.close()

        p_fill = self._forward_fill_headers(p_raw)
        s_fill = self._forward_fill_headers(s_raw)
        idx = picked_col - 1
        if idx < 0 or idx >= max_col:
            raise ValueError("所选列超出范围")

        g_primary = p_fill[idx]
        g_secondary = s_fill[idx]

        if looks_like_group_header(g_primary):
            group_name = g_primary
            group_row = "primary"
        elif looks_like_group_header(g_secondary):
            group_name = g_secondary
            group_row = "secondary"
        else:
            group_name = g_primary or g_secondary
            group_row = "primary" if g_primary else "secondary"

        if not group_name:
            raise ValueError("未识别到分组头，请检查主/次表头行设置")

        if group_row == "primary":
            group_cols = [c for c in range(1, max_col + 1) if p_fill[c - 1] == group_name]
            item_raw = s_raw
            backup_raw = p_raw
        else:
            group_cols = [c for c in range(1, max_col + 1) if s_fill[c - 1] == group_name]
            item_raw = p_raw
            backup_raw = s_raw

        direct_col = None
        sum_cols = []
        item_names = []

        for c in group_cols:
            item_name = normalize_text(item_raw[c - 1]) or normalize_text(backup_raw[c - 1])
            if looks_like_subtotal(item_name):
                direct_col = c
                continue
            if item_name and item_name != group_name:
                sum_cols.append(c)
                item_names.append(item_name)

        if direct_col is None:
            for c in group_cols:
                p = normalize_text(p_raw[c - 1])
                s = normalize_text(s_raw[c - 1])
                if looks_like_subtotal(p) or looks_like_subtotal(s):
                    direct_col = c
                    break

        if not sum_cols:
            fallback_cols = [c for c in group_cols if c != direct_col]
            for c in fallback_cols:
                label = normalize_text(item_raw[c - 1]) or normalize_text(backup_raw[c - 1]) or f"列{get_column_letter(c)}"
                sum_cols.append(c)
                item_names.append(label)

        return {
            "group_name": group_name,
            "direct_col": get_column_letter(direct_col) if direct_col else "",
            "sum_cols": [get_column_letter(c) for c in sum_cols],
            "item_names": item_names,
        }

    def auto_detect_all_dimensions(self, dims, header_row_primary: int, header_row_secondary: int):
        excel_path_text = self.excel_path_var.get().strip()
        sheet_name = self.sheet_var.get().strip()
        excel_path = Path(excel_path_text)
        if not excel_path.exists():
            raise FileNotFoundError(f"找不到表格文件: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        max_col = min(ws.max_column, 200)
        p_raw = [normalize_text(ws.cell(row=header_row_primary, column=c).value) for c in range(1, max_col + 1)]
        s_raw = [normalize_text(ws.cell(row=header_row_secondary, column=c).value) for c in range(1, max_col + 1)]
        wb.close()

        p_fill = self._forward_fill_headers(p_raw)
        s_fill = self._forward_fill_headers(s_raw)

        def compact(v):
            return normalize_text(v).replace(" ", "").replace("（", "(").replace("）", ")").lower()

        new_dims = []
        for d in dims:
            dim_name = str(d.get("name", "")).strip()
            if not dim_name:
                continue
            cdim = compact(dim_name)

            group_cols = []
            for c in range(1, max_col + 1):
                if cdim and (cdim in compact(p_fill[c - 1]) or cdim in compact(s_fill[c - 1])):
                    group_cols.append(c)

            if not group_cols:
                new_dims.append(d)
                continue

            direct_col = ""
            sum_cols = []
            item_names = []

            for c in group_cols:
                p = p_raw[c - 1]
                s = s_raw[c - 1]
                label = pick_recommended_header(p, s)
                if looks_like_subtotal(p) or looks_like_subtotal(s) or looks_like_subtotal(label):
                    if not direct_col:
                        direct_col = get_column_letter(c)
                    continue

                if label and compact(label) != cdim:
                    sum_cols.append(get_column_letter(c))
                    item_names.append(label)

            if not sum_cols and group_cols:
                fallback = [c for c in group_cols if get_column_letter(c) != direct_col]
                if not fallback:
                    fallback = group_cols
                sum_cols = [get_column_letter(c) for c in fallback]
                if not item_names:
                    item_names = [p_raw[c - 1] or s_raw[c - 1] or f"列{get_column_letter(c)}" for c in fallback]

            new_dims.append(
                {
                    "name": dim_name,
                    "direct_col": direct_col,
                    "sum_cols": sum_cols,
                    "item_names": item_names,
                }
            )

        return new_dims

    def build_mapping_validation_messages(self, raw_mapping: dict):
        messages = []
        try:
            parsed = validate_mapping_config(raw_mapping)
        except Exception as exc:
            return [f"配置错误：{exc}"]

        used_direct = {}
        used_sum = {}
        for dim in parsed["dimensions"]:
            dname = dim["name"]
            direct_col = dim["direct_col"]
            if direct_col:
                used_direct.setdefault(direct_col, []).append(dname)
            for c in dim["sum_cols"]:
                used_sum.setdefault(c, []).append(dname)

            if direct_col and direct_col in dim["sum_cols"]:
                messages.append(f"[{dname}] 总分列与子项列重复（{get_column_letter(direct_col)}）")

        for col, names in used_direct.items():
            if len(names) > 1:
                messages.append(f"总分列重复：{get_column_letter(col)} 被 {', '.join(names)} 共用")

        for col, names in used_sum.items():
            if len(names) > 1:
                messages.append(f"子项列重复：{get_column_letter(col)} 被 {', '.join(names)} 共用")

        if not messages:
            messages.append("校验通过：未发现明显配置问题。")
        return messages

    def pick_column_from_header(self, parent, header_row_primary: int, header_row_secondary: int):
        excel_path_text = self.excel_path_var.get().strip()
        sheet_name = self.sheet_var.get().strip()
        if not excel_path_text or not sheet_name:
            messagebox.showwarning("提示", "请先在主界面选择 Excel 和工作表。", parent=parent)
            return None, None

        excel_path = Path(excel_path_text)
        if not excel_path.exists():
            messagebox.showerror("错误", f"找不到表格文件：\n{excel_path}", parent=parent)
            return None, None

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        max_col = min(ws.max_column, 120)

        items = []
        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            p = normalize_text(ws.cell(row=header_row_primary, column=c).value)
            s = normalize_text(ws.cell(row=header_row_secondary, column=c).value)
            label = pick_recommended_header(p, s)
            if not p and not s:
                continue
            items.append((col_letter, label, p, s))
        wb.close()

        result = {"col": None, "name": None}
        dialog = tk.Toplevel(parent)
        dialog.title("点选列号")
        dialog.transient(parent)
        dialog.grab_set()
        dialog.geometry("680x460")

        ttk.Label(dialog, text="点击一行即可选择该列").pack(anchor="w", padx=12, pady=(10, 6))

        table = ttk.Treeview(dialog, columns=("col", "label", "p", "s"), show="headings", height=16)
        table.heading("col", text="列号")
        table.heading("label", text="推荐表头")
        table.heading("p", text=f"主表头({header_row_primary})")
        table.heading("s", text=f"次表头({header_row_secondary})")
        table.column("col", width=70, anchor="center")
        table.column("label", width=180, anchor="w")
        table.column("p", width=180, anchor="w")
        table.column("s", width=180, anchor="w")
        table.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        for i, (col_letter, label, p, s) in enumerate(items):
            table.insert("", "end", iid=str(i), values=(col_letter, label, p, s))

        def confirm():
            sel = table.selection()
            if not sel:
                return
            idx = int(sel[0])
            result["col"] = items[idx][0]
            result["name"] = items[idx][1]
            dialog.destroy()

        table.bind("<Double-1>", lambda _e: confirm())
        btn_row = ttk.Frame(dialog)
        btn_row.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(btn_row, text="取消", command=dialog.destroy).pack(side="right")
        ttk.Button(btn_row, text="选择", command=confirm).pack(side="right", padx=(0, 8))

        self.wait_window(dialog)
        return result["col"], result["name"]

    def open_dimension_editor(self, parent, existing=None, header_row_primary=4, header_row_secondary=5):
        result = {"value": None}
        dialog = tk.Toplevel(parent)
        dialog.title("维度配置")
        dialog.transient(parent)
        dialog.grab_set()
        dialog.resizable(False, False)

        existing = existing or {"name": "", "direct_col": "", "sum_cols": [], "item_names": []}
        name_var = tk.StringVar(value=str(existing.get("name", "")))
        direct_col_var = tk.StringVar(value=str(existing.get("direct_col", "")))
        sum_cols_var = tk.StringVar(value=",".join(str(x) for x in existing.get("sum_cols", [])))
        item_names_var = tk.StringVar(value=",".join(str(x) for x in existing.get("item_names", [])))

        ttk.Label(dialog, text="维度名称").grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))
        ttk.Entry(dialog, textvariable=name_var, width=32).grid(row=0, column=1, padx=12, pady=(12, 4))
        ttk.Button(
            dialog,
            text="按分组自动填充",
            command=lambda: self._auto_fill_dimension_by_group(
                dialog,
                name_var,
                direct_col_var,
                sum_cols_var,
                item_names_var,
                header_row_primary,
                header_row_secondary,
            ),
        ).grid(row=0, column=2, padx=(0, 12), pady=(12, 4))

        ttk.Label(dialog, text="总分列").grid(row=1, column=0, sticky="w", padx=12, pady=4)
        ttk.Entry(dialog, textvariable=direct_col_var, width=32).grid(row=1, column=1, padx=12, pady=4)
        ttk.Button(
            dialog,
            text="点选总分列",
            command=lambda: self._pick_direct_col(
                dialog,
                direct_col_var,
                header_row_primary,
                header_row_secondary,
            ),
        ).grid(row=1, column=2, padx=(0, 12), pady=4)

        ttk.Label(dialog, text="子项列号").grid(row=2, column=0, sticky="w", padx=12, pady=4)
        ttk.Entry(dialog, textvariable=sum_cols_var, width=32).grid(row=2, column=1, padx=12, pady=4)
        ttk.Button(
            dialog,
            text="点选子项列",
            command=lambda: self._pick_item_col(
                dialog,
                sum_cols_var,
                item_names_var,
                header_row_primary,
                header_row_secondary,
            ),
        ).grid(row=2, column=2, padx=(0, 12), pady=4)
        ttk.Label(dialog, text="示例：H,I,J 或 8,9,10", foreground="#666").grid(row=3, column=1, sticky="w", padx=12, pady=(0, 4))

        ttk.Label(dialog, text="子项表头").grid(row=4, column=0, sticky="w", padx=12, pady=4)
        ttk.Entry(dialog, textvariable=item_names_var, width=32).grid(row=4, column=1, padx=12, pady=4)
        ttk.Label(dialog, text="逗号分隔，可留空自动读表头", foreground="#666").grid(row=5, column=1, sticky="w", padx=12, pady=(0, 8))

        def on_confirm():
            result["value"] = {
                "name": name_var.get().strip(),
                "direct_col": direct_col_var.get().strip(),
                "sum_cols": self.split_csv(sum_cols_var.get()),
                "item_names": self.split_csv(item_names_var.get()),
            }
            dialog.destroy()

        btn_row = ttk.Frame(dialog)
        btn_row.grid(row=6, column=0, columnspan=2, pady=(0, 12))
        ttk.Button(btn_row, text="取消", command=dialog.destroy).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btn_row, text="确定", command=on_confirm).grid(row=0, column=1)

        self.wait_window(dialog)
        return result["value"]

    def _pick_direct_col(self, dialog, direct_col_var, header_row_primary, header_row_secondary):
        col, _name = self.pick_column_from_header(dialog, header_row_primary, header_row_secondary)
        if col:
            direct_col_var.set(col)

    def _pick_item_col(self, dialog, sum_cols_var, item_names_var, header_row_primary, header_row_secondary):
        col, name = self.pick_column_from_header(dialog, header_row_primary, header_row_secondary)
        if not col:
            return
        cols = self.split_csv(sum_cols_var.get())
        if col not in cols:
            cols.append(col)
            sum_cols_var.set(",".join(cols))
        if name:
            names = self.split_csv(item_names_var.get())
            if len(names) < len(cols):
                names.append(name)
            item_names_var.set(",".join(names))

    def _auto_fill_dimension_by_group(
        self,
        dialog,
        name_var,
        direct_col_var,
        sum_cols_var,
        item_names_var,
        header_row_primary,
        header_row_secondary,
    ):
        picked_col, _name = self.pick_column_from_header(dialog, header_row_primary, header_row_secondary)
        if not picked_col:
            return
        try:
            mapping = self.detect_group_mapping_by_column(header_row_primary, header_row_secondary, picked_col)
            if not name_var.get().strip():
                name_var.set(mapping["group_name"])
            direct_col_var.set(mapping["direct_col"])
            sum_cols_var.set(",".join(mapping["sum_cols"]))
            item_names_var.set(",".join(mapping["item_names"]))
        except Exception as exc:
            messagebox.showerror("自动填充失败", str(exc), parent=dialog)

    def open_mapping_editor(self):
        dialog = tk.Toplevel(self)
        dialog.title("配置映射")
        dialog.transient(self)
        dialog.grab_set()
        dialog.geometry("900x700")

        cfg = self.mapping_config_raw
        name_var = tk.StringVar(value=str(cfg.get("name", "自定义模板")))
        name_col_var = tk.StringVar(value=str(cfg.get("name_col", 2)))
        data_start_row_var = tk.StringVar(value=str(cfg.get("data_start_row", 6)))
        header_primary_var = tk.StringVar(value=str(cfg.get("header_row_primary", 4)))
        header_secondary_var = tk.StringVar(value=str(cfg.get("header_row_secondary", 5)))
        score_max_var = tk.StringVar(value=str(cfg.get("score_max", 10)))
        dims = [dict(d) for d in cfg.get("dimensions", [])]

        root = ttk.Frame(dialog, padding=12)
        root.pack(fill="both", expand=True)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(3, weight=1)

        ttk.Label(root, text="模板名").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(root, textvariable=name_var).grid(row=0, column=1, sticky="ew", pady=4)

        top_grid = ttk.Frame(root)
        top_grid.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(4, 6))
        for i in range(6):
            top_grid.columnconfigure(i, weight=1 if i % 2 else 0)

        ttk.Label(top_grid, text="姓名列").grid(row=0, column=0, sticky="w")
        ttk.Entry(top_grid, textvariable=name_col_var, width=10).grid(row=0, column=1, sticky="w", padx=(6, 18))
        ttk.Label(top_grid, text="数据起始行").grid(row=0, column=2, sticky="w")
        ttk.Entry(top_grid, textvariable=data_start_row_var, width=10).grid(row=0, column=3, sticky="w", padx=(6, 18))
        ttk.Label(top_grid, text="score_max").grid(row=0, column=4, sticky="w")
        ttk.Entry(top_grid, textvariable=score_max_var, width=10).grid(row=0, column=5, sticky="w", padx=(6, 0))

        ttk.Label(top_grid, text="主表头行").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(top_grid, textvariable=header_primary_var, width=10).grid(row=1, column=1, sticky="w", padx=(6, 18), pady=(6, 0))
        ttk.Label(top_grid, text="次表头行").grid(row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(top_grid, textvariable=header_secondary_var, width=10).grid(row=1, column=3, sticky="w", padx=(6, 18), pady=(6, 0))

        ttk.Label(root, text="维度配置（表头 + 列号）", font=("Microsoft YaHei", 10, "bold")).grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 4))

        table = ttk.Treeview(root, columns=("name", "direct_col", "sum_cols", "item_names"), show="headings", height=11)
        table.heading("name", text="维度名")
        table.heading("direct_col", text="总分列")
        table.heading("sum_cols", text="子项列号")
        table.heading("item_names", text="子项表头")
        table.column("name", width=120, anchor="w")
        table.column("direct_col", width=80, anchor="center")
        table.column("sum_cols", width=180, anchor="w")
        table.column("item_names", width=280, anchor="w")
        table.grid(row=3, column=0, columnspan=2, sticky="nsew")
        ttk.Label(root, text="提示：按住左键拖动行可调整维度顺序", foreground="#666").grid(
            row=7, column=0, columnspan=2, sticky="w", pady=(6, 0)
        )
        
        validate_frame = ttk.LabelFrame(root, text="映射校验")
        validate_frame.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        validate_frame.columnconfigure(0, weight=1)
        validation_text = tk.Text(validate_frame, height=6, wrap="word", relief="flat")
        validation_text.grid(row=0, column=0, sticky="ew", padx=8, pady=8)
        validation_text.configure(state="disabled")
        dim_by_iid = {}
        drag_state = {"iid": None, "moved": False, "start_y": 0}

        def build_current_raw():
            return {
                "name": name_var.get().strip() or "自定义模板",
                "name_col": name_col_var.get().strip(),
                "data_start_row": data_start_row_var.get().strip(),
                "header_row_primary": header_primary_var.get().strip(),
                "header_row_secondary": header_secondary_var.get().strip(),
                "score_max": score_max_var.get().strip(),
                "dimensions": dims,
            }

        def refresh_validation():
            messages = self.build_mapping_validation_messages(build_current_raw())
            validation_text.configure(state="normal")
            validation_text.delete("1.0", "end")
            for msg in messages:
                validation_text.insert("end", f"- {msg}\n")
            validation_text.configure(state="disabled")

        def sync_dims_from_table():
            dims[:] = [dim_by_iid[iid] for iid in table.get_children("")]

        def refresh_table():
            dim_by_iid.clear()
            for row_id in table.get_children():
                table.delete(row_id)
            for i, d in enumerate(dims):
                iid = f"dim_{i}"
                dim_by_iid[iid] = d
                table.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(
                        d.get("name", ""),
                        d.get("direct_col", ""),
                        ",".join(str(x) for x in d.get("sum_cols", [])),
                        ",".join(str(x) for x in d.get("item_names", [])),
                    ),
                )
            refresh_validation()

        refresh_table()

        action = ttk.Frame(root)
        action.grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 6))

        def add_dim():
            value = self.open_dimension_editor(
                dialog,
                header_row_primary=int(header_primary_var.get().strip() or "4"),
                header_row_secondary=int(header_secondary_var.get().strip() or "5"),
            )
            if value:
                dims.append(value)
                refresh_table()

        def edit_dim():
            sel = table.selection()
            if not sel:
                return
            iid = sel[0]
            value = self.open_dimension_editor(
                dialog,
                dict(dim_by_iid[iid]),
                header_row_primary=int(header_primary_var.get().strip() or "4"),
                header_row_secondary=int(header_secondary_var.get().strip() or "5"),
            )
            if value:
                dim_by_iid[iid] = value
                sync_dims_from_table()
                refresh_table()

        def del_dim():
            sel = table.selection()
            if not sel:
                return
            for iid in sel:
                if iid in dim_by_iid:
                    del dim_by_iid[iid]
                table.delete(iid)
            sync_dims_from_table()
            refresh_table()

        def on_drag_start(event):
            iid = table.identify_row(event.y)
            drag_state["iid"] = iid if iid else None
            drag_state["moved"] = False
            drag_state["start_y"] = event.y
            if iid:
                table.selection_set(iid)

        def on_drag_motion(event):
            moving = drag_state["iid"]
            if not moving:
                return
            if abs(event.y - drag_state["start_y"]) < 4:
                return
            target = table.identify_row(event.y)
            if not target or target == moving:
                return
            table.move(moving, "", table.index(target))
            drag_state["moved"] = True

        def on_drag_end(_event):
            if not drag_state["iid"]:
                return
            if drag_state["moved"]:
                sync_dims_from_table()
                refresh_table()
            drag_state["iid"] = None
            drag_state["moved"] = False

        table.bind("<ButtonPress-1>", on_drag_start)
        table.bind("<B1-Motion>", on_drag_motion)
        table.bind("<ButtonRelease-1>", on_drag_end)

        ttk.Button(action, text="新增维度", command=add_dim).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(action, text="编辑维度", command=edit_dim).grid(row=0, column=1, padx=(0, 6))
        ttk.Button(action, text="删除维度", command=del_dim).grid(row=0, column=2, padx=(0, 6))

        def detect_all():
            try:
                hp = int(header_primary_var.get().strip() or "4")
                hs = int(header_secondary_var.get().strip() or "5")
                new_dims = self.auto_detect_all_dimensions(dims, hp, hs)
                dims[:] = new_dims
                refresh_table()
            except Exception as exc:
                messagebox.showerror("识别失败", str(exc), parent=dialog)

        ttk.Button(action, text="一键识别全部", command=detect_all).grid(row=0, column=3)

        btn_row = ttk.Frame(root)
        btn_row.grid(row=6, column=0, columnspan=2, sticky="e", pady=(10, 0))

        def apply_and_close():
            try:
                raw = {
                    "name": name_var.get().strip() or "自定义模板",
                    "name_col": name_col_var.get().strip(),
                    "data_start_row": int(data_start_row_var.get().strip()),
                    "header_row_primary": int(header_primary_var.get().strip()),
                    "header_row_secondary": int(header_secondary_var.get().strip()),
                    "score_max": float(score_max_var.get().strip()),
                    "dimensions": dims,
                }
                self.apply_mapping_config(raw)
                if not self.current_template_path:
                    filename = f"{sanitize_template_filename(raw['name'])}.json"
                    self.current_template_path = self.template_dir / filename
                self.autosave_current_template()
                self.reload_people()
                dialog.destroy()
            except Exception as exc:
                messagebox.showerror("配置错误", str(exc), parent=dialog)

        ttk.Button(btn_row, text="取消", command=dialog.destroy).pack(side="right")
        ttk.Button(btn_row, text="应用", command=apply_and_close).pack(side="right", padx=(0, 8))
        refresh_validation()

    def save_template(self):
        save_path = filedialog.asksaveasfilename(
            title="保存模板",
            defaultextension=".json",
            initialdir=str(self.template_dir),
            initialfile=f"{self.mapping_config_raw.get('name', 'mapping')}.json",
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
        )
        if not save_path:
            return
        try:
            save_path_obj = Path(save_path)
            self.write_template_file(save_path_obj, self.mapping_config_raw)
            self.current_template_path = save_path_obj
            self.template_dir = save_path_obj.parent
            self.template_file_var.set(f"模板文件：{self.current_template_path}")
            self.save_app_settings()
            messagebox.showinfo("保存成功", f"模板已保存：\n{save_path}")
        except Exception as exc:
            messagebox.showerror("保存失败", f"保存模板失败：\n{exc}")

    def load_template(self):
        file_path = filedialog.askopenfilename(
            title="加载模板",
            initialdir=str(self.template_dir),
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
        )
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                raw = json.load(f)
            self.apply_mapping_config(raw)
            self.current_template_path = Path(file_path)
            self.template_dir = self.current_template_path.parent
            self.template_file_var.set(f"模板文件：{self.current_template_path}")
            self.save_app_settings()
            self.reload_people()
            messagebox.showinfo("加载成功", f"模板已加载：\n{file_path}")
        except Exception as exc:
            messagebox.showerror("加载失败", f"加载模板失败：\n{exc}")

    def set_excel_file(self, excel_path: Path):
        self.excel_path_var.set(str(excel_path))
        self.populate_sheets(excel_path)
        self.reload_people()

    def populate_sheets(self, excel_path: Path):
        try:
            sheet_names = list_sheet_names(excel_path)
        except Exception as exc:
            messagebox.showerror("错误", f"读取工作表失败:\n{exc}")
            self.sheet_combo["values"] = []
            self.sheet_var.set("")
            return

        self.sheet_combo["values"] = sheet_names
        target = DEFAULT_SHEET_NAME if DEFAULT_SHEET_NAME in sheet_names else sheet_names[0]
        self.sheet_var.set(target)

    def on_choose_excel(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            initialdir=str(BASE_DIR),
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if not file_path:
            return
        self.set_excel_file(Path(file_path))

    def on_sheet_change(self, _event=None):
        self.reload_people()

    def reload_people(self):
        excel_path_text = self.excel_path_var.get().strip()
        sheet_name = self.sheet_var.get().strip()
        if not excel_path_text or not sheet_name:
            return

        excel_path = Path(excel_path_text)
        if not excel_path.exists():
            messagebox.showerror("错误", f"找不到表格文件:\n{excel_path}")
            self.clear_view()
            return

        try:
            people = load_people_scores(excel_path, sheet_name, self.mapping_config)
        except Exception as exc:
            messagebox.showerror("错误", f"读取数据失败:\n{exc}")
            self.clear_view()
            return

        self.people = people
        self.people_by_name = {p["name"]: p for p in people}
        self.search_var.set("")
        self.refresh_name_list([p["name"] for p in self.people])
        if self.export_batch_button is not None:
            self.export_batch_button.config(state="normal" if self.people else "disabled")

        if self.filtered_names:
            self.listbox.selection_set(0)
            self.listbox.event_generate("<<ListboxSelect>>")
        else:
            messagebox.showwarning("提示", "没有读取到员工数据，请检查工作表配置。")
            self.clear_chart_and_details()

    def clear_chart_and_details(self):
        self.current_person = None
        self.dimension_text_map = {}
        self.ax.clear()
        self.canvas.draw_idle()
        if self.export_current_button is not None:
            self.export_current_button.config(state="disabled")
        if self.export_batch_button is not None:
            self.export_batch_button.config(state="normal" if self.people else "disabled")
        self.detail_title_var.set("当前维度：")
        for row_id in self.detail_table.get_children():
            self.detail_table.delete(row_id)

    def clear_view(self):
        self.people = []
        self.people_by_name = {}
        self.refresh_name_list([])
        self.clear_chart_and_details()

    def refresh_name_list(self, names):
        self.filtered_names = names
        self.listbox.delete(0, tk.END)
        for name in names:
            self.listbox.insert(tk.END, name)

    def on_search(self, _event=None):
        keyword = normalize_search_text(self.search_var.get().strip())
        if not keyword:
            names = [p["name"] for p in self.people]
        else:
            names = [
                p["name"]
                for p in self.people
                if keyword in normalize_search_text(p["name"])
            ]
        self.refresh_name_list(names)
        if names:
            self.listbox.selection_set(0)
            self.listbox.event_generate("<<ListboxSelect>>")
        else:
            self.clear_chart_and_details()

    def on_select(self, _event=None):
        sel = self.listbox.curselection()
        if not sel:
            return
        name = self.listbox.get(sel[0])
        person = self.people_by_name.get(name)
        if not person:
            return

        self.current_person = person
        tick_texts = draw_radar(self.ax, name, self.labels, person["scores"], self.mapping_config["score_max"])
        self.dimension_text_map = {txt: txt.get_text().replace("\n", "").strip() for txt in tick_texts}
        if self.export_current_button is not None:
            self.export_current_button.config(state="normal")
        self.update_detail_panel()
        self.canvas.draw_idle()

    def export_current_chart(self):
        if not self.current_person:
            messagebox.showwarning("提示", "请先选择员工后再导出。")
            return

        person_name = normalize_export_name(self.current_person["name"])
        default_name = f"{person_name}_能力雷达图V6.png"
        self.export_person_chart(self.current_person, default_name)

    def ask_people_for_bulk_export(self):
        if not self.people:
            return []

        result = {"names": []}
        dialog = tk.Toplevel(self)
        dialog.title("批量导出 - 选择人员")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        ttk.Label(dialog, text="请选择要批量导出的人员（勾选即可）：").grid(
            row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(12, 6)
        )

        names = [p["name"] for p in self.people]
        checked_map = {n: tk.BooleanVar(value=False) for n in names}
        if self.current_person and self.current_person["name"] in checked_map:
            checked_map[self.current_person["name"]].set(True)

        list_frame = ttk.Frame(dialog)
        list_frame.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=12)
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        canvas = tk.Canvas(list_frame, width=280, height=280, highlightthickness=0)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scrollbar.set)

        check_frame = ttk.Frame(canvas)
        window_id = canvas.create_window((0, 0), window=check_frame, anchor="nw")

        def on_frame_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            canvas.itemconfigure(window_id, width=event.width)

        check_frame.bind("<Configure>", on_frame_configure)
        canvas.bind("<Configure>", on_canvas_configure)

        for idx, n in enumerate(names):
            ttk.Checkbutton(check_frame, text=f"{n}", variable=checked_map[n]).grid(
                row=idx, column=0, sticky="w", padx=4, pady=2
            )

        def select_all():
            for var in checked_map.values():
                var.set(True)

        def clear_sel():
            for var in checked_map.values():
                var.set(False)

        def confirm():
            result["names"] = [n for n, var in checked_map.items() if var.get()]
            dialog.destroy()

        ttk.Button(dialog, text="全选", command=select_all).grid(row=2, column=0, padx=(12, 6), pady=(8, 10), sticky="w")
        ttk.Button(dialog, text="清空", command=clear_sel).grid(row=2, column=1, padx=6, pady=(8, 10), sticky="w")
        btn_row = ttk.Frame(dialog)
        btn_row.grid(row=2, column=2, padx=(6, 12), pady=(8, 10), sticky="e")
        ttk.Button(btn_row, text="取消", command=dialog.destroy).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btn_row, text="导出", command=confirm).grid(row=0, column=1)

        self.wait_window(dialog)
        return result["names"]

    def export_multiple_people_charts(self):
        if not self.people:
            messagebox.showwarning("提示", "当前没有可导出的人员数据。")
            return

        selected_names = self.ask_people_for_bulk_export()
        if not selected_names:
            return

        output_dir = filedialog.askdirectory(
            title="选择批量导出目录",
            initialdir=str(BASE_DIR),
        )
        if not output_dir:
            return

        output_path = Path(output_dir)
        success_count = 0
        failed = []

        for name in selected_names:
            person = self.people_by_name.get(name)
            if not person:
                failed.append(name)
                continue
            filename = f"{normalize_export_name(name)}_能力雷达图V6.png"
            save_path = output_path / filename
            try:
                self.save_person_chart(person, save_path)
                success_count += 1
            except Exception:
                failed.append(name)

        if failed:
            messagebox.showwarning(
                "批量导出完成",
                f"成功导出 {success_count} 张，失败 {len(failed)} 人：\n" + "、".join(failed),
            )
        else:
            messagebox.showinfo("批量导出完成", f"已成功导出 {success_count} 张图片到：\n{output_path}")

    def export_person_chart(self, person: dict, default_name: str):
        save_path = filedialog.asksaveasfilename(
            title="导出雷达图",
            defaultextension=".png",
            initialdir=str(BASE_DIR),
            initialfile=default_name,
            filetypes=[("PNG 图片", "*.png"), ("JPEG 图片", "*.jpg"), ("所有文件", "*.*")],
        )
        if not save_path:
            return

        try:
            self.save_person_chart(person, Path(save_path))
            messagebox.showinfo("导出成功", f"已导出图片：\n{save_path}")
        except Exception as exc:
            messagebox.showerror("导出失败", f"导出图片失败：\n{exc}")

    def save_person_chart(self, person: dict, save_path: Path):
        export_fig = plt.Figure(
            figsize=(7.2, 7.2),
            dpi=220,
            facecolor="#FAFBFC",
            constrained_layout=True,
        )
        export_ax = export_fig.add_subplot(111, polar=True)
        export_fig.set_constrained_layout_pads(w_pad=0.03, h_pad=0.03, hspace=0.02, wspace=0.02)
        draw_radar(export_ax, person["name"], self.labels, person["scores"], self.mapping_config["score_max"])
        export_fig.savefig(save_path, dpi=220, bbox_inches="tight", facecolor=export_fig.get_facecolor())
        plt.close(export_fig)

    def on_chart_pick(self, event):
        artist = event.artist
        if artist not in self.dimension_text_map:
            return
        self.set_dimension(self.dimension_text_map[artist])

    def set_dimension(self, dim_name: str):
        self.current_dimension = dim_name
        self.update_detail_panel()

    def update_detail_panel(self):
        if not self.current_person:
            return
        if self.current_dimension not in self.labels and self.labels:
            self.current_dimension = self.labels[0]

        detail = self.current_person["details"].get(self.current_dimension)
        if not detail:
            return

        self.detail_title_var.set(f"当前维度：{self.current_dimension}")

        for row_id in self.detail_table.get_children():
            self.detail_table.delete(row_id)

        for item in detail["items"]:
            self.detail_table.insert("", "end", values=(item["label"], f"{item['value']:.1f}"))

        self.detail_table.insert("", "end", values=("小计", f"{detail['subtotal']:.1f}"))

