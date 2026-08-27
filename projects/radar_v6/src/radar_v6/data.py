from pathlib import Path

import openpyxl


def to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return 0.0
        try:
            return float(stripped)
        except ValueError:
            return 0.0
    return 0.0

def read_cell_value(ws_data, row, col):
    return ws_data.cell(row=row, column=col).value

def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").strip()

def normalize_search_text(value: str) -> str:
    # 搜索时忽略姓名中的全角/半角空格
    return "".join(str(value).replace("\u3000", " ").split()).lower()

def normalize_export_name(name: str) -> str:
    bad_chars = '<>:"/\\|?*'
    cleaned = str(name).replace("\u3000", " ").strip()
    for ch in bad_chars:
        cleaned = cleaned.replace(ch, "_")
    return "".join(cleaned.split()) or "未命名"

def get_col_label(ws_data, col: int, fallback: str, header_row_primary: int, header_row_secondary: int) -> str:
    primary = normalize_text(read_cell_value(ws_data, header_row_primary, col))
    if primary:
        return primary
    secondary = normalize_text(read_cell_value(ws_data, header_row_secondary, col))
    if secondary:
        return secondary
    return fallback

def get_dimension_detail(ws_data, row: int, dim: dict, header_row_primary: int, header_row_secondary: int):
    direct_col = dim.get("direct_col")
    direct_val = read_cell_value(ws_data, row, direct_col) if direct_col else None
    use_direct = (direct_col is not None) and (direct_val not in (None, ""))

    items = []
    if direct_col is not None and len(dim["sum_cols"]) == 1 and dim["sum_cols"][0] == direct_col:
        label = f"{dim['name']}总分"
        items.append({"label": label, "value": to_number(direct_val)})
        subtotal = to_number(direct_val)
        source = "小计列"
    else:
        subtotal = 0.0
        for idx, col in enumerate(dim["sum_cols"]):
            if dim.get("item_names") and idx < len(dim["item_names"]) and str(dim["item_names"][idx]).strip():
                label = str(dim["item_names"][idx]).strip()
            else:
                label = get_col_label(ws_data, col, f"列{col}", header_row_primary, header_row_secondary)
            val = to_number(read_cell_value(ws_data, row, col))
            subtotal += val
            items.append({"label": label, "value": val})

        if use_direct:
            subtotal = to_number(direct_val)
            source = "小计列"
        else:
            source = "子项求和"

    return {
        "items": items,
        "subtotal": subtotal,
        "source": source,
    }

def load_people_scores(excel_path: Path, sheet_name: str, mapping_cfg: dict):
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    ws_data = wb_data[sheet_name]

    people = []
    row = mapping_cfg["data_start_row"]
    while row <= ws_data.max_row:
        name = ws_data.cell(row=row, column=mapping_cfg["name_col"]).value
        if name is None or str(name).strip() == "":
            row += 1
            continue

        scores = []
        details = {}
        for dim in mapping_cfg["dimensions"]:
            detail = get_dimension_detail(
                ws_data,
                row,
                dim,
                mapping_cfg["header_row_primary"],
                mapping_cfg["header_row_secondary"],
            )
            details[dim["name"]] = detail
            scores.append(detail["subtotal"])

        people.append({
            "name": str(name).strip(),
            "scores": scores,
            "details": details,
        })
        row += 1

    return people

def list_sheet_names(excel_path: Path):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names

